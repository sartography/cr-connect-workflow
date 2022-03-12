import json

from tests.base_test import BaseTest

from crc import session
from crc.models.task_log import TaskLogModel, TaskLogQuery, TaskLogQuerySchema
from crc.models.user import UserModel

from openpyxl import load_workbook
from io import BytesIO


class TestDownloadLogsForStudy(BaseTest):
    @staticmethod
    def add_log(study_id, workflow_id, task, workflow_spec_id, log_data):
        task_log = TaskLogModel(level=log_data['level'],
                                code=log_data['code'],
                                message=log_data['message'],
                                study_id=study_id,
                                workflow_id=workflow_id,
                                task=task,
                                user_uid='joe',
                                workflow_spec_id=workflow_spec_id)
        session.add(task_log)
        session.commit()

    def test_download_logs_for_study(self):
        workflow = self.create_workflow('empty_workflow')
        workflow_api = self.get_workflow_api(workflow)
        task = workflow_api.next_task
        study_id = workflow.study_id

        log_data = {'level': 'metrics',
                    'code': 'test_code',
                    'message': 'This is a message.'}
        self.add_log(study_id, workflow.id, task.name, 'empty_workflow', log_data)
        log_data = {'level': 'metrics',
                    'code': 'another_test_code',
                    'message': 'This is another message.'}
        self.add_log(study_id, workflow.id, task.name, 'empty_workflow', log_data)
        log_data = {'level': 'metrics',
                    'code': 'a_third_test_code',
                    'message': 'This is a third message.'}
        self.add_log(study_id, workflow.id, task.name, 'empty_workflow', log_data)

        # Run the query, which should include a 'download_url' link that we can click on.
        url = f'/v1.0/study/{workflow.study_id}/log'
        task_log_query = TaskLogQuery()
        user = session.query(UserModel).filter_by(uid=self.test_uid).first()
        rv = self.app.put(url, headers=self.logged_in_headers(user), content_type="application/json",
                          data=TaskLogQuerySchema().dump(task_log_query))
        self.assert_success(rv)
        log_query = json.loads(rv.get_data(as_text=True))
        self.assertIsNotNone(log_query['download_url'])

        # Use the provided link to get the file.
        rv = self.app.get(log_query['download_url'])
        self.assert_success(rv)
        wb = load_workbook(BytesIO(rv.data))
        ws = wb.active

        self.assertEqual(4, ws.max_row)
        self.assertEqual('Category', ws['A1'].value)
        self.assertEqual('empty_workflow', ws['B2'].value)
        self.assertEqual('metrics', ws['C3'].value)
        self.assertEqual('a_third_test_code', ws['D4'].value)
