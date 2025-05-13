from openpyxl import Workbook, load_workbook
from tempfile import NamedTemporaryFile

from typing import List

from crc.services.reference_file_service import ReferenceFileService


class SpreadsheetService(object):

    @staticmethod
    def create_spreadsheet(data: List[dict], headers: List[str] = None, title: str = None):
        """The length of headers must be the same as the number of items in the dictionaries,
           and the order must match up.
           The title is used for the worksheet, not the filename."""

        wb = Workbook(write_only=True)
        ws = wb.create_sheet()
        if title:
            ws.title = title
        if headers:
            ws.append(headers)
        for row in data:
            ws.append(list(row.values()))

        with NamedTemporaryFile() as tmp:
            wb.save(tmp.name)
            tmp.seek(0)
            stream = tmp.read()
            return stream


    def create_enum_dict_from_spreadsheet(self, name):
        """Creates a dictionary from a spreadsheet.
           Meant for form enums.
           We assume 2 columns, the first column is the key and the second column is the value."""
        path_name = ReferenceFileService.file_path(name)
        wb = load_workbook(filename=path_name)
        ws = wb.active
        data_dict = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            data_dict[row[0]] = row[1]
        return data_dict
